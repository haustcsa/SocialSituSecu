import openpyxl
import os
import argparse
import time
import torch
import torch.optim as optim
import pandas as pd
# from torch.utils.tensorboard import SummaryWriter
import datetime
from Mao_ConvNet.my_dataset import MyDataSet
from Mao_ConvNet.models.base_model import BaseModel
from Mao_ConvNet.utils import read_split_data, create_lr_scheduler, get_params_groups, train_one_epoch, evaluate, \
    img_transform
from PIL import ImageFile

ImageFile.LOAD_TRUNCATED_IMAGES = True


def main(args, modelname):
    device = torch.device(args.device if torch.cuda.is_available() else "cpu")
    # device = torch.device("cpu")
    print(f"using {device} device.")

    if os.path.exists("./w/" + modelname) is False:
        os.makedirs("./w/" + modelname)
    SummaryWriter_path = "./run/" + modelname + time.strftime("%Y%m%d%H")
    if os.path.exists(SummaryWriter_path) is False:
        os.makedirs(SummaryWriter_path)
    # tb_writer = SummaryWriter(SummaryWriter_path + '/logs')

    train_images_path, train_images_label, val_images_path, val_images_label = read_split_data(args.data_path)

    data_transform = img_transform(modelname)

    # 实例化训练数据集
    train_dataset = MyDataSet(images_path=train_images_path,
                              images_class=train_images_label,
                              transform=data_transform["train"])

    # 实例化验证数据集
    val_dataset = MyDataSet(images_path=val_images_path,
                            images_class=val_images_label,
                            transform=data_transform["val"])

    batch_size = args.batch_size
    nw = min([os.cpu_count(), batch_size if batch_size > 1 else 0, 8])  # number of workers
    print('Using {} dataloader workers every process'.format(nw))
    train_loader = torch.utils.data.DataLoader(train_dataset,
                                               batch_size=batch_size,
                                               shuffle=True,
                                               pin_memory=True,
                                               num_workers=nw,
                                               collate_fn=train_dataset.collate_fn)

    val_loader = torch.utils.data.DataLoader(val_dataset,
                                             batch_size=batch_size,
                                             shuffle=False,
                                             pin_memory=True,
                                             num_workers=nw,
                                             collate_fn=val_dataset.collate_fn)

    model = BaseModel(num_classes=args.num_classes, name=modelname).to(device)
    if args.weights != "":
        assert os.path.exists(args.weights), "weights file: '{}' not exist.".format(args.weights)
        weights_dict = torch.load(args.weights, map_location=device)["model"]
        # 删除有关分类类别的权重
        for k in list(weights_dict.keys()):
            if "head" in k:
                del weights_dict[k]
        print(model.load_state_dict(weights_dict, strict=False))

    if args.freeze_layers:
        for name, para in model.named_parameters():
            # 除head外，其他权重全部冻结
            if "head" not in name:
                para.requires_grad_(False)
            else:
                print("training {}".format(name))

    # pg = [p for p in model.parameters() if p.requires_grad]
    pg = get_params_groups(model, weight_decay=args.wd)
    optimizer = optim.AdamW(pg, lr=args.lr, weight_decay=args.wd)
    lr_scheduler = create_lr_scheduler(optimizer, len(train_loader), args.epochs,
                                       warmup=True, warmup_epochs=1)

    best_acc = 0.00
    # start = time.time()

    record_train_loss = []
    record_train_acc = []
    record_val_loss = []
    record_val_acc = []
    record_learning_rate = []

    for epoch in range(args.epochs):
        # train
        train_loss, train_acc = train_one_epoch(model=model,
                                                optimizer=optimizer,
                                                data_loader=train_loader,
                                                device=device,
                                                epoch=epoch,
                                                lr_scheduler=lr_scheduler)

        # validate
        val_loss, val_acc = evaluate(model=model,
                                     data_loader=val_loader,
                                     device=device,
                                     epoch=epoch)

        tags = ["train_loss", "train_acc", "val_loss", "val_acc", "learning_rate"]
        # tb_writer.add_scalar(tags[0], train_loss, epoch)
        # tb_writer.add_scalar(tags[1], train_acc, epoch)
        # tb_writer.add_scalar(tags[2], val_loss, epoch)
        # tb_writer.add_scalar(tags[3], val_acc, epoch)
        # tb_writer.add_scalar(tags[4], optimizer.param_groups[0]["lr"], epoch)
        record_train_loss.append(train_loss)
        record_train_acc.append(train_acc)
        record_val_loss.append(val_loss)
        record_val_acc.append(val_acc)
        record_learning_rate.append(optimizer.param_groups[0]["lr"])
        if best_acc < val_acc:
            save_path = "./w/" + modelname + "/" + modelname + "_" + str(epoch) + "_" + str(best_acc) + ".pth"
            torch.save(model.state_dict(), save_path)
            best_acc = val_acc
        assert len(record_train_loss) == len(record_train_acc) == len(record_val_loss) == len(record_val_acc) == len(record_learning_rate),\
            f"{epoch}轮Record记录不一致"
    # end = time.time()
    # print("Training 耗时为:{:.1f}".format(end - start))
    return record_train_loss, record_train_acc, record_val_loss, record_val_acc, record_learning_rate


def write_excel(i, modelname, e_path, dataset):
    wb = openpyxl.load_workbook(e_path)
    ws = wb.active
    ws.cell(row=1, column=i, value=modelname)
    for epoch, data in enumerate(dataset):
        # 指定行列给单元格赋值
        epoch +=2
        ws.cell(row=epoch, column=i, value=str(data))
    # 保存
    wb.save(e_path)
    wb.close()


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--num_classes', type=int, default=4)
    parser.add_argument('--epochs', type=int, default=400)
    parser.add_argument('--batch-size', type=int, default=24)
    parser.add_argument('--lr', type=float, default=5e-4)
    parser.add_argument('--wd', type=float, default=5e-2)
    # 数据集所在根目录
    parser.add_argument('--data-path', type=str, default="./dataset")
    # 预训练权重路径，如果不想载入就设置为空字符
    parser.add_argument('--weights', type=str, default='', help='initial weights path')
    # 是否冻结head以外所有权重
    parser.add_argument('--freeze-layers', type=bool, default=False)
    parser.add_argument('--device', default='cuda:0', help='device id (i.e. 0 or 0,1 or cpu)')
    opt = parser.parse_args()

    # AlexNet,VggNet,GoogleNet,ResNet,DenseNet,MobileNet_v2,MobileNet_v3,ShuffleNet_v1,ShuffleNet_v2,EfficientNet_v1_B0(B0-B7),EfficientNetV2_s(s,m,l),GhostNet,ConvNext
    # main(opt, modelname='ResNet')
    # main(opt, modelname='DenseNet')
    # model_list = ['MobileNet_v3', 'ShuffleNet_v1', 'ShuffleNet_v2', 'EfficientNet_v1_B0','EfficientNetV2_s', 'GhostNet', 'ConvNext']
    model_list = ['MobileNet_v3','MobileNetV3_Ghost','MobileNetV3_Ghost221SE','MobileNetV3_Ghost112SE']
    record_files = ['./record/record_tl.xlsx', './record/record_ta.xlsx', './record/record_vl.xlsx', './record/record_va.xlsx', './record/record_lr.xlsx']
    for record_file in record_files:
        if not os.path.exists(record_file):
            df = pd.DataFrame()
            df.to_excel(record_file)
    n = int(pd.read_excel(record_files[0]).shape[1])

    for i, model_name in enumerate(model_list):
        i += 1+n
        print(f"第{i}个模型:{model_name}开始训练...")
        record_tl, record_ta, record_vl, record_va, record_lr = main(opt, modelname=model_name)
        write_excel(i, model_name, './record/record_tl.xlsx', record_tl)
        write_excel(i, model_name, './record/record_ta.xlsx', record_ta)
        write_excel(i, model_name, './record/record_vl.xlsx', record_vl)
        write_excel(i, model_name, './record/record_va.xlsx', record_va)
        write_excel(i, model_name, './record/record_lr.xlsx', record_lr)