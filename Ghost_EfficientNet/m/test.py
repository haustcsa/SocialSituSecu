# dicts = {
#     "social": ['social_keyword*', 'corr_social_keyword*', 'corr_social_freq*'],
#     "fake": ['fake_keyword*', 'corr_fake_keyword*', 'corr_fake_freq*'],
#     "covid": ['covid_keyword*', 'corr_covid_keyword*', 'corr_covid_freq*'],
# }
# times = ['1', '3', '7']
#
# sql = "select top 50* from("
# for w_i, w in enumerate(dicts):
#     for i in range(3):
#         sql += "select top 50 " + dicts[w][1].replace('*', times[i]) + " as word," + dicts[w][2].replace('*', times[
#             i]) + " as req from tb_WordCloudCorr_Visual " \
#                   "where (" + dicts[w][0].replace('*', times[i]) + ")   = '" + "俄罗斯" + "' "
#         if w_i == 2 and i == 2:
#             pass
#         else:
#             sql += " union all "
# sql += ")a"
# print(sql)
import os
import pandas as pd
import time

# import random
#
# frames = [random.randint(0, 10) for i in range(1000)]
# for i, x in enumerate(frames):
#     time.sleep(0.01)
#     jindu = int(int(i + 1) / len(frames) * 100)
#     print(f"\r第{i + 1}帧添加成功,进度：" + (jindu) * "*" + (100 - jindu - 1) * "-" + str(jindu) + "%", end='')
# aaa = {"a":"a1","b":"b1","c":"c1"}
# for i ,a in enumerate(aaa):
#     print(i,aaa[a])
import openpyxl


def write_excel(i, modelname, e_path, dataset):
    wb = openpyxl.load_workbook(e_path)
    ws = wb.active
    ws.cell(row=1, column=i, value=modelname)
    for epoch, data in enumerate(dataset):
        # 指定行列给单元格赋值
        epoch +=2
        ws.cell(row=epoch, column=i, value=str(data))
    # 保存
    wb.save(e_path)
    wb.close()

